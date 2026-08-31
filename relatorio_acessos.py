"""
Mostra quem está acessando o site e com que frequência, usando os dados
gravados em 'historico_acessos' (ver services/auth.py -> registrar_acesso).

IMPORTANTE: só existem dados a partir do momento em que a tabela
'historico_acessos' foi criada (rode criar_tabela_historico_acessos.py
uma vez, se ainda não rodou). Logins anteriores a isso não ficaram
registrados, pois o site não gravava esse histórico até agora.

Uso:
    python relatorio_acessos.py

Isso gera também um arquivo 'relatorio_acessos.xlsx' na mesma pasta, com o
resumo formatado (cabeçalho em negrito, colunas de data formatadas e largura
ajustada automaticamente).
"""
import pandas as pd
from sqlalchemy import text
from services.database import obter_engine
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

pd.set_option("display.max_rows", None)
pd.set_option("display.width", 120)
ARQUIVO_SAIDA = "relatorio_acessos.xlsx"


def salvar_planilha(resumo: pd.DataFrame, caminho: str = ARQUIVO_SAIDA) -> None:
    """Salva o resumo em .xlsx com formatação básica (fonte, cabeçalho,
    largura de colunas e formato de data)."""
    resumo.to_excel(caminho, sheet_name="Resumo de acessos", index=False)

    wb = load_workbook(caminho)
    ws = wb["Resumo de acessos"]

    fonte_padrao = "Arial"
    colunas_data = {"primeiro_acesso", "ultimo_acesso"}

    # Cabeçalho em negrito
    for col_idx, col_nome in enumerate(resumo.columns, start=1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = Font(name=fonte_padrao, bold=True)

    # Formata células de dados (fonte + formato de data quando aplicável)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for celula in row:
            col_nome = resumo.columns[celula.column - 1]
            if col_nome in colunas_data:
                celula.number_format = "dd/mm/yyyy hh:mm"
            celula.font = Font(name=fonte_padrao)

    # Ajusta largura das colunas ao conteúdo
    for col_idx, col_nome in enumerate(resumo.columns, start=1):
        letra = get_column_letter(col_idx)
        maior = max(
            [len(str(col_nome))]
            + [len(str(v)) for v in resumo[col_nome].astype(str).tolist()]
        )
        ws.column_dimensions[letra].width = min(max(maior + 2, 10), 40)

    ws.freeze_panes = "A2"
    wb.save(caminho)


def main():
    engine = obter_engine()
    with engine.connect() as conn:
        # Resumo por usuário: total de acessos, primeiro e último acesso,
        # e quantos acessos nos últimos 7 e 30 dias (indicador de uso "regular")
        resumo = pd.read_sql(
            text("""
                SELECT
                    u.login,
                    u.nome,
                    u.ativo,
                    COUNT(h.id) AS total_acessos,
                    MIN(h.acessado_em) AS primeiro_acesso,
                    MAX(h.acessado_em) AS ultimo_acesso,
                    COUNT(h.id) FILTER (WHERE h.acessado_em >= NOW() - INTERVAL '7 days')  AS acessos_7d,
                    COUNT(h.id) FILTER (WHERE h.acessado_em >= NOW() - INTERVAL '30 days') AS acessos_30d
                FROM usuarios u
                LEFT JOIN historico_acessos h ON h.usuario_id = u.id
                GROUP BY u.login, u.nome, u.ativo
                ORDER BY ultimo_acesso DESC NULLS LAST
            """),
            conn,
        )

    print("\n=== Resumo de acessos por usuário ===\n")
    if resumo.empty:
        print("Nenhum usuário cadastrado.")
    else:
        print(resumo.to_string(index=False))

    print(
        "\nDica: 'acessos_7d' e 'acessos_30d' são bons indicadores de uso "
        "regular. Usuários com 'total_acessos' = 0 nunca fizeram login "
        "(ou só logaram antes da tabela de histórico existir)."
    )

    salvar_planilha(resumo)
    print(f"\nPlanilha salva em: {ARQUIVO_SAIDA}")


if __name__ == "__main__":
    main()
